"""显式触发的 AI 两层识别任务。"""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

from parser_poc.ark_adapter import ArkStructuredAdapter
from parser_poc.extraction import extract_validated_table
from parser_poc.orchestration import run_xlsx_sheet_with_adapter

from ..infrastructure.db import update_processing_job
from ..settings import AI_MAX_SHEETS, AI_OUTLINE_HARD_TOKENS, AI_OUTLINE_TARGET_TOKENS, AI_PROFILE


def run_ai_recognition(job_id: str, source_path: str) -> None:
    """按配置的 Sheet 上限执行 Layer 1/Layer 2，并只保存状态摘要。"""
    try:
        update_processing_job(job_id, status="running", phase="连接 AI Provider", progress_percent=5)
        adapter = ArkStructuredAdapter.from_environment(AI_PROFILE)  # type: ignore[arg-type]
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        try:
            sheet_names = workbook.sheetnames[:AI_MAX_SHEETS]
        finally:
            workbook.close()
        results = []
        for index, sheet_name in enumerate(sheet_names, 1):
            progress = 10 + int((index - 1) / max(len(sheet_names), 1) * 80)
            update_processing_job(job_id, status="running", phase=f"AI 识别 Sheet：{sheet_name}", progress_percent=progress)

            def report_sheet_phase(phase: str) -> None:
                """把耗时的本地读取与模型等待区别显示，避免前端只看到固定 10%。"""
                update_processing_job(
                    job_id,
                    status="running",
                    phase=f"Sheet {sheet_name}：{phase}",
                    progress_percent=min(progress + 5, 89),
                )

            outlines, details, validations = run_xlsx_sheet_with_adapter(
                path=Path(source_path),
                sheet_name=sheet_name,
                adapter=adapter,
                target_tokens=AI_OUTLINE_TARGET_TOKENS,
                hard_tokens=AI_OUTLINE_HARD_TOKENS,
                progress_callback=report_sheet_phase,
            )
            report_sheet_phase("Python 校验与提取")
            extracted_tables = []
            validation_index = 0
            for detail_response in details:
                for proposal in detail_response.proposals:
                    validation = validations[validation_index]
                    validation_index += 1
                    if validation.outcome.value not in {"accepted", "adjusted"}:
                        continue
                    extracted_tables.append(
                        extract_validated_table(
                            path=Path(source_path),
                            proposal=proposal,
                            validation=validation,
                            extracted_table_id=f"{job_id}_{index}_{len(extracted_tables) + 1}",
                            metric_type="unknown",
                        )
                    )
            results.append({
                "sheet_name": sheet_name,
                "outline_response_count": len(outlines),
                "detail_response_count": len(details),
                "outline_responses": [response.model_dump(mode="json") for response in outlines],
                "boundary_proposals": [
                    proposal.model_dump(mode="json")
                    for response in details
                    for proposal in response.proposals
                ],
                "boundary_validations": [validation.model_dump(mode="json") for validation in validations],
                "validation_outcome_counts": {
                    outcome.value: sum(1 for item in validations if item.outcome == outcome)
                    for outcome in {item.outcome for item in validations}
                },
                # 只有 Python 对源文件回读并通过边界校验后，才生成提取数据。
                "extracted_tables": extracted_tables,
            })
        update_processing_job(job_id, status="completed", phase="AI 两层识别完成", progress_percent=100, result={"sheets": results, "provider": AI_PROFILE, "max_sheets": AI_MAX_SHEETS})
    except Exception as exc:
        update_processing_job(job_id, status="failed", phase="AI 识别失败", progress_percent=100, error_message=str(exc))
