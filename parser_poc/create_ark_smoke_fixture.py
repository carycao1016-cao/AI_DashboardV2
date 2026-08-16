"""从真实 PoC Tab 复制一个低成本 Ark Smoke workbook。

默认截取 Quantum 的第一个 Golden 物理表，保留值、公式、数字格式、合并单元格和基础尺寸，
不修改原始工作簿。Smoke fixture 只用于验证 Provider 连通性和结构化输出，不用于替代完整 Golden。
"""

from __future__ import annotations

import argparse
from copy import copy
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries


def create_fixture(source: Path, output: Path, *, sheet_name: str, max_row: int, max_column: int) -> None:
    source_book = load_workbook(source, data_only=False)
    if sheet_name not in source_book.sheetnames:
        raise ValueError("Sheet not found: %s" % sheet_name)
    source_sheet = source_book[sheet_name]
    if max_row < 1 or max_column < 1:
        raise ValueError("Fixture bounds must be positive")

    fixture_book = load_workbook(source, data_only=False)
    # 删除原工作簿中的全部 Sheet，再创建唯一 Smoke Sheet，避免把其他表带入请求。
    for name in list(fixture_book.sheetnames):
        del fixture_book[name]
    target_sheet = fixture_book.create_sheet(sheet_name)

    for row in source_sheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_column):
        for source_cell in row:
            target_cell = target_sheet[source_cell.coordinate]
            target_cell.value = source_cell.value
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format
            if source_cell.alignment:
                target_cell.alignment = copy(source_cell.alignment)
            if source_cell.protection:
                target_cell.protection = copy(source_cell.protection)

    for reference in source_sheet.merged_cells.ranges:
        min_col, min_row, source_max_col, source_max_row = range_boundaries(str(reference))
        if min_row >= 1 and source_max_row <= max_row and min_col >= 1 and source_max_col <= max_column:
            target_sheet.merge_cells(str(reference))
    for index in range(1, max_column + 1):
        letter = get_column_letter(index)
        if letter in source_sheet.column_dimensions:
            target_sheet.column_dimensions[letter] = copy(source_sheet.column_dimensions[letter])
    for index in range(1, max_row + 1):
        if index in source_sheet.row_dimensions:
            target_sheet.row_dimensions[index] = copy(source_sheet.row_dimensions[index])
    target_sheet.freeze_panes = source_sheet.freeze_panes
    target_sheet.sheet_view.showGridLines = source_sheet.sheet_view.showGridLines
    output.parent.mkdir(parents=True, exist_ok=True)
    fixture_book.save(output)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--source",
        type=Path,
        default=Path(__file__).parents[1] / "PoC/Quantum Tab/Tabs_%95.xlsx",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path(__file__).parents[1] / "outputs/ark_smoke/Tabs_%95_first_table.xlsx",
    )
    parser.add_argument("--sheet", default="ban1_%Sig")
    parser.add_argument("--max-row", type=int, default=78)
    parser.add_argument("--max-column", type=int, default=28)
    args = parser.parse_args()
    create_fixture(
        args.source,
        args.output,
        sheet_name=args.sheet,
        max_row=args.max_row,
        max_column=args.max_column,
    )
    print("已生成 Ark Smoke fixture：%s" % args.output)


if __name__ == "__main__":
    main()
