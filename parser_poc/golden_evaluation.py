"""Golden annotation loading and status-only evaluation for table recognition."""

from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path
from typing import Iterable, Optional, Sequence

from openpyxl import load_workbook
from openpyxl.utils import range_boundaries
from pydantic import BaseModel, ConfigDict, Field, model_validator

from parser_poc.contracts import (
    BoundaryValidationResult,
    CoarseRange,
    SignificanceLayout,
    ValidationOutcome,
)


def parse_row_spec(value: object) -> list[int]:
    """Parse Golden row coordinates such as 5, 5:7 or 5,7:9 without text inference."""
    if value is None or value == "":
        return []
    if isinstance(value, int):
        return [value]
    rows: list[int] = []
    for part in str(value).split(","):
        start_end = [int(item.strip()) for item in part.split(":")]
        if len(start_end) == 1:
            rows.append(start_end[0])
        elif len(start_end) == 2 and start_end[0] <= start_end[1]:
            rows.extend(range(start_end[0], start_end[1] + 1))
        else:
            raise ValueError("Invalid Golden row specification: %s" % value)
    return rows


class GoldenTableAnnotation(BaseModel):
    """Only structural Golden fields required for a status-only report."""

    model_config = ConfigDict(extra="forbid")

    table_id: str = Field(min_length=1)
    sheet_name: str = Field(min_length=1)
    table_range: str = Field(min_length=1)
    title_rows: list[int] = Field(default_factory=list)
    header_rows: list[int] = Field(min_length=1)
    base_rows: list[int] = Field(default_factory=list)
    data_rows: list[int] = Field(min_length=1)
    footnote_rows: list[int] = Field(default_factory=list)
    expected_outline_status: str
    expected_validation_result: ValidationOutcome
    header_depth: int = Field(ge=1)
    has_explicit_base: bool
    significance_layout: SignificanceLayout
    source_family: str = "unknown"

    @model_validator(mode="after")
    def validate_rows_within_range(self) -> "GoldenTableAnnotation":
        _min_column, min_row, _max_column, max_row = range_boundaries(self.table_range)
        all_regions = [
            *self.title_rows,
            *self.header_rows,
            *self.base_rows,
            *self.data_rows,
            *self.footnote_rows,
        ]
        if any(row < min_row or row > max_row for row in all_regions):
            raise ValueError("Annotated region rows must be inside table_range")
        return self


def _as_boolean(value: object) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"true", "1", "yes", "y"}


def load_golden_annotations(path: Path) -> list[GoldenTableAnnotation]:
    """Read only Table_Annotation fields; business text is intentionally ignored."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    info_sheet = workbook["Workbook_Info"]
    info_rows = info_sheet.iter_rows(values_only=True)
    info_headers = next(info_rows)
    info_positions = {str(name): index for index, name in enumerate(info_headers) if name}
    source_families = {
        str(row[info_positions["Golden_ID"]]): str(row[info_positions["Source_Type"]])
        for row in info_rows
        if row[info_positions["Golden_ID"]]
    }
    sheet = workbook["Table_Annotation"]
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows)
    positions = {str(name): index for index, name in enumerate(headers) if name}
    required = {
        "Golden_ID", "Table_ID", "Sheet_Name", "Table_Range", "Title_Rows", "Header_Rows",
        "Base_Rows", "Data_Rows", "Footnote_Rows",
        "Significance_Layout", "Expected_Outline_Status", "Expected_Validation_Result",
        "Header_Depth", "Has_Explicit_Base",
    }
    missing = required.difference(positions)
    if missing:
        raise ValueError("Golden template is missing columns: %s" % ", ".join(sorted(missing)))
    annotations = []
    for row in rows:
        table_id = row[positions["Table_ID"]]
        if not table_id:
            continue
        annotations.append(
            GoldenTableAnnotation(
                table_id=str(table_id),
                sheet_name=str(row[positions["Sheet_Name"]]),
                table_range=str(row[positions["Table_Range"]]),
                title_rows=parse_row_spec(row[positions["Title_Rows"]]),
                header_rows=parse_row_spec(row[positions["Header_Rows"]]),
                base_rows=parse_row_spec(row[positions["Base_Rows"]]),
                data_rows=parse_row_spec(row[positions["Data_Rows"]]),
                footnote_rows=parse_row_spec(row[positions["Footnote_Rows"]]),
                significance_layout=str(row[positions["Significance_Layout"]]),
                expected_outline_status=str(row[positions["Expected_Outline_Status"]]),
                expected_validation_result=str(row[positions["Expected_Validation_Result"]]),
                header_depth=int(row[positions["Header_Depth"]]),
                has_explicit_base=_as_boolean(row[positions["Has_Explicit_Base"]]),
                source_family=source_families.get(str(row[positions["Golden_ID"]]), "unknown"),
            )
        )
    return annotations


def outline_covers_annotation(annotation: GoldenTableAnnotation, candidates: Iterable[CoarseRange]) -> bool:
    """Coverage means Header or first data row hit, or full table containment."""
    _min_col, table_start, _max_col, table_end = range_boundaries(annotation.table_range)
    first_data_row = min(annotation.data_rows)
    for candidate in candidates:
        header_hit = any(candidate.start_row <= row <= candidate.end_row for row in annotation.header_rows)
        first_data_hit = candidate.start_row <= first_data_row <= candidate.end_row
        contains_table = candidate.start_row <= table_start and candidate.end_row >= table_end
        if header_hit or first_data_hit or contains_table:
            return True
    return False


def evaluate_golden(
    annotations: Sequence[GoldenTableAnnotation],
    candidates_by_sheet: dict[str, Sequence[CoarseRange]],
    validations_by_table_id: dict[str, BoundaryValidationResult],
) -> dict[str, object]:
    """Produce aggregate statuses only; labels, values and question text are excluded."""
    total = len(annotations)
    covered = [annotation for annotation in annotations if outline_covers_annotation(annotation, candidates_by_sheet.get(annotation.sheet_name, []))]
    outcomes = Counter(
        result.outcome.value for result in validations_by_table_id.values()
    )
    breakdown: dict[str, Counter[str]] = defaultdict(Counter)
    for annotation in annotations:
        scope = "source_family=%s|sheet=%s|header_depth=%s|significance_layout=%s" % (
            annotation.source_family,
            annotation.sheet_name,
            annotation.header_depth,
            annotation.significance_layout.value,
        )
        breakdown[scope]["total"] += 1
        if outline_covers_annotation(annotation, candidates_by_sheet.get(annotation.sheet_name, [])):
            breakdown[scope]["outline_covered"] += 1
        result = validations_by_table_id.get(annotation.table_id)
        if result:
            breakdown[scope]["validation_%s" % result.outcome.value] += 1

    return {
        "schema_name": "golden_table_recognition_report",
        "schema_version": "0.1.0-poc",
        "table_count": total,
        "outline_covered_table_count": len(covered),
        "outline_missed_table_count": total - len(covered),
        "outline_coverage_rate": 0 if not total else len(covered) / total,
        "validation_outcome_distribution": dict(sorted(outcomes.items())),
        "breakdowns": {key: dict(sorted(value.items())) for key, value in sorted(breakdown.items())},
    }
