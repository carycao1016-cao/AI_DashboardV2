"""Offline two-layer orchestration shell with a controlled structured-output Stub."""

from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import Any, Callable, Iterable, Sequence

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries
from pydantic import BaseModel

from parser_poc.contracts import (
    BoundaryValidationResult,
    CoarseRange,
    DetailWindowRequest,
    DetailWindowResponse,
    OutlineStatus,
    SheetOutlineResponse,
    StructuredGenerationAdapter,
    TableBoundaryProposal,
    ValidationCategory,
    ValidationOutcome,
)
from parser_poc.workbook_scan import (
    DEFAULT_HARD_TOKENS,
    DEFAULT_TARGET_TOKENS,
    build_detail_window,
    cache_read_only_sheet_window,
    plan_detail_windows,
    scan_workbook,
    xlsx_sheet_metadata,
)


class ControlledStubAdapter:
    """Returns explicit fixture responses only; no semantic inference is performed."""

    def __init__(self, responses: Iterable[BaseModel]) -> None:
        self._responses = list(responses)
        self.calls: list[dict[str, Any]] = []

    def generate_structured(self, *, task_name: str, payload: dict[str, Any], output_model: type[BaseModel]) -> BaseModel:
        self.calls.append({"task_name": task_name, "payload": payload, "output_model": output_model.__name__})
        for index, response in enumerate(self._responses):
            if isinstance(response, output_model):
                return self._responses.pop(index)
        raise LookupError("No controlled Stub response matches %s" % output_model.__name__)


def validate_proposal(proposal: TableBoundaryProposal, *, total_rows: int, total_columns: int) -> BoundaryValidationResult:
    """Validate only physical bounds and declared region rows in this offline shell."""
    categories: list[ValidationCategory] = []
    try:
        min_column, min_row, max_column, max_row = range_boundaries(proposal.source_range)
    except ValueError:
        categories.append(ValidationCategory.RANGE_INVALID)
    else:
        if min_row < 1 or max_row > total_rows or min_column < 1 or max_column > total_columns:
            categories.append(ValidationCategory.RANGE_INVALID)
        region_rows = [
            *proposal.regions.title_rows,
            *proposal.regions.header_rows,
            *proposal.regions.base_rows,
            *proposal.regions.data_rows,
            *proposal.regions.footnote_rows,
            *proposal.regions.significance_locations,
        ]
        if any(row < min_row or row > max_row for row in region_rows):
            categories.append(ValidationCategory.RANGE_INVALID)
        if not proposal.regions.header_rows or not proposal.regions.data_rows:
            categories.append(ValidationCategory.STRUCTURE_UNRESOLVED)

    return BoundaryValidationResult(
        proposal_id=proposal.proposal_id,
        outcome=ValidationOutcome.REVIEW_REQUIRED if categories else ValidationOutcome.ACCEPTED,
        categories=categories,
    )


def normalize_edge_rows(proposal: TableBoundaryProposal, value_sheet: Any) -> tuple[TableBoundaryProposal, list[str]]:
    """移除明确的分页/空白边缘行，并保留可审计的修正说明。

    AI 可能会把候选窗口边缘的 ``#page`` 或空白行纳入 source_range。
    这些行不能作为表内容，但表内的空白行不能据此删除，因为 Quantum
    的显著性标记可能位于数据行之后。故这里只检查范围最外侧，并且只
    在该行没有被 proposal 的任何结构区域声明时修正。
    """
    min_col, min_row, max_col, max_row = range_boundaries(proposal.source_range)
    declared_rows = {
        *proposal.regions.title_rows,
        *proposal.regions.header_rows,
        *proposal.regions.base_rows,
        *proposal.regions.data_rows,
        *proposal.regions.footnote_rows,
        *proposal.regions.significance_locations,
    }

    def edge_kind(row_number: int) -> str:
        values = [value_sheet.cell(row_number, column).value for column in range(min_col, max_col + 1)]
        non_empty = [str(value).strip() for value in values if value not in (None, "") and str(value).strip()]
        if not non_empty:
            return "blank"
        if non_empty[0] == "#page":
            return "page_marker"
        return "content"

    new_min, new_max = min_row, max_row
    corrections: list[str] = []
    while new_min <= new_max and new_min not in declared_rows and edge_kind(new_min) in {"blank", "page_marker"}:
        corrections.append("trimmed_top_%s_row_%s" % (edge_kind(new_min), new_min))
        new_min += 1
    while new_max >= new_min and new_max not in declared_rows and edge_kind(new_max) in {"blank", "page_marker"}:
        corrections.append("trimmed_bottom_%s_row_%s" % (edge_kind(new_max), new_max))
        new_max -= 1
    if not corrections:
        return proposal, corrections
    adjusted = proposal.model_copy(
        update={
            "source_range": "%s%s:%s%s"
            % (get_column_letter(min_col), new_min, get_column_letter(max_col), new_max)
        }
    )
    return adjusted, corrections


def normalize_known_summary_rows(proposal: TableBoundaryProposal, value_sheet: Any) -> tuple[TableBoundaryProposal, list[str]]:
    """把有明确源标签的 Sigma 汇总行从数据区归入脚注区。

    该规则有意保持很窄：只接受首列标签精确为 ``Sigma`` 的行，避免把
    普通选项文本或未知表格样式错误重分类。它解决的是模型在 Quantum
    物理范围正确时，仍把 Tab 软件的汇总统计行误列为业务数据的问题。
    """
    min_col, _min_row, _max_col, _max_row = range_boundaries(proposal.source_range)
    summary_rows = []
    for row_number in proposal.regions.data_rows:
        label = value_sheet.cell(row_number, min_col).value
        normalized_label = str(label).strip().casefold() if label not in (None, "") else ""
        if normalized_label == "sigma":
            summary_rows.append(row_number)
    if not summary_rows:
        return proposal, []
    regions = proposal.regions.model_copy(
        update={
            "data_rows": [row for row in proposal.regions.data_rows if row not in summary_rows],
            "footnote_rows": sorted({*proposal.regions.footnote_rows, *summary_rows}),
        }
    )
    normalized = proposal.model_copy(update={"regions": regions})
    return normalized, ["reclassified_sigma_row_%s_as_footnote" % row for row in summary_rows]


def _candidate_ranges(responses: Sequence[SheetOutlineResponse]) -> list[tuple[str, int, int]]:
    return [
        (candidate.candidate_id, candidate.start_row, candidate.end_row)
        for response in responses
        for candidate in response.candidates
        if candidate.status in {OutlineStatus.COMPLETE, OutlineStatus.AMBIGUOUS, OutlineStatus.NEEDS_MORE_CONTEXT}
    ]


def build_detail_requests(
    *,
    sheet_name: str,
    total_rows: int,
    outline_responses: Sequence[SheetOutlineResponse],
    context_before: int = 20,
    context_after: int = 20,
    max_candidate_gap_rows: int = 20,
) -> list[DetailWindowRequest]:
    """Group positional candidates without merging their identities."""
    candidates = _candidate_ranges(outline_responses)
    windows = plan_detail_windows(
        [(start, end) for _identifier, start, end in candidates],
        total_rows,
        context_before,
        context_after,
        max_candidate_gap_rows,
    )
    identifiers_by_range: dict[tuple[int, int], list[str]] = defaultdict(list)
    for identifier, start, end in candidates:
        identifiers_by_range[(start, end)].append(identifier)

    requests = []
    for window in windows:
        candidate_ids = []
        for item in window["candidate_ranges"]:
            start, end = (int(value) for value in item.split(":"))
            candidate_ids.extend(identifiers_by_range[(start, end)])
        requests.append(
            DetailWindowRequest(
                sheet_name=sheet_name,
                window_id=window["window_id"],
                candidate_ids=candidate_ids,
                window_row_start=window["window_row_start"],
                window_row_end=window["window_row_end"],
                payload=window,
            )
        )
    return requests


def run_sheet_with_adapter(
    *,
    sheet_name: str,
    total_rows: int,
    total_columns: int,
    outline_chunks: Sequence[dict[str, Any]],
    adapter: StructuredGenerationAdapter,
) -> tuple[list[SheetOutlineResponse], list[DetailWindowResponse], list[BoundaryValidationResult]]:
    """Run the contracts with an adapter; source-value extraction is intentionally absent."""
    outline_responses = [
        adapter.generate_structured(
            task_name="sheet_outline",
            payload={"sheet_name": sheet_name, **chunk},
            output_model=SheetOutlineResponse,
        )
        for chunk in outline_chunks
    ]
    detail_requests = build_detail_requests(
        sheet_name=sheet_name,
        total_rows=total_rows,
        outline_responses=outline_responses,
    )
    detail_responses = [
        adapter.generate_structured(
            task_name="detail_window",
            payload=request.model_dump(mode="json"),
            output_model=DetailWindowResponse,
        )
        for request in detail_requests
    ]
    validations = [
        validate_proposal(proposal, total_rows=total_rows, total_columns=total_columns)
        for response in detail_responses
        for proposal in response.proposals
    ]
    return outline_responses, detail_responses, validations


def run_xlsx_sheet_with_adapter(
    *,
    path: Path,
    sheet_name: str,
    adapter: StructuredGenerationAdapter,
    target_tokens: int = DEFAULT_TARGET_TOKENS,
    hard_tokens: int = DEFAULT_HARD_TOKENS,
    detail_context_before: int = 20,
    detail_context_after: int = 20,
    max_candidate_gap_rows: int = 20,
    progress_callback: Callable[[str], None] | None = None,
) -> tuple[list[SheetOutlineResponse], list[DetailWindowResponse], list[BoundaryValidationResult]]:
    """对一个 XLSX Sheet 执行真实的两层流程，并由 Python 提供 Detail 样本。

    Layer 1 只接收 WorkbookScanSummary 的 Outline。只有模型返回候选后，Python
    才按候选窗口重读原始单元格并生成 Detail Window；因此模型不会直接读取整张 Sheet。
    """
    if progress_callback:
        progress_callback("构建结构摘要")
    summary = scan_workbook(
        path,
        target_tokens=target_tokens,
        hard_tokens=hard_tokens,
        sheet_name=sheet_name,
    )
    sheet_summary = summary["sheets"][0]
    if progress_callback:
        progress_callback("等待 AI Outline 响应")
    outline_responses = []
    for chunk in sheet_summary["outline_chunks"]:
        outline_responses.append(adapter.generate_structured(
            task_name="sheet_outline",
            payload={"sheet_name": sheet_name, **chunk},
            output_model=SheetOutlineResponse,
        ))
    total_rows = range_boundaries(sheet_summary["scan_range"])[3]
    total_columns = range_boundaries(sheet_summary["scan_range"])[2]
    detail_requests = build_detail_requests(
        sheet_name=sheet_name,
        total_rows=total_rows,
        outline_responses=outline_responses,
        context_before=detail_context_before,
        context_after=detail_context_after,
        max_candidate_gap_rows=max_candidate_gap_rows,
    )
    if not detail_requests:
        return outline_responses, [], []

    if progress_callback:
        progress_callback("读取候选 Detail 窗口")
    layout = xlsx_sheet_metadata(path)[sheet_name]
    value_book = load_workbook(path, read_only=True, data_only=True)
    formula_book = load_workbook(path, read_only=True, data_only=False)
    try:
        # 后续校正会多次按坐标读取。仅缓存 AI 已提出候选的 Detail 窗口，
        # 避免 ReadOnlyWorksheet.cell() 为每一个坐标重复解析整份 Sheet XML。
        detail_start = min(request.window_row_start for request in detail_requests)
        detail_end = max(request.window_row_end for request in detail_requests)
        value_sheet, formula_sheet = cache_read_only_sheet_window(
            value_book[sheet_name],
            formula_book[sheet_name],
            min_row=detail_start,
            max_row=detail_end,
        )
        detail_responses = []
        if progress_callback:
            progress_callback("等待 AI Detail 响应")
        for request in detail_requests:
            detail_payload = build_detail_window(
                value_sheet,
                formula_sheet,
                layout,
                request.model_dump(mode="json")["payload"],
                target_tokens,
                hard_tokens,
            )
            detail_responses.append(
                adapter.generate_structured(
                    task_name="detail_window",
                    payload={
                        "sheet_name": sheet_name,
                        "window_id": request.window_id,
                        "candidate_ids": request.candidate_ids,
                        **detail_payload,
                    },
                    output_model=DetailWindowResponse,
                )
            )
        if progress_callback:
            progress_callback("校正候选边界")
        normalized_responses: list[DetailWindowResponse] = []
        normalization_notes: dict[str, list[str]] = {}
        for response in detail_responses:
            normalized_proposals = []
            for proposal in response.proposals:
                normalized, corrections = normalize_edge_rows(proposal, value_sheet)
                normalized, summary_corrections = normalize_known_summary_rows(normalized, value_sheet)
                corrections.extend(summary_corrections)
                normalized_proposals.append(normalized)
                if corrections:
                    normalization_notes[normalized.proposal_id] = corrections
            normalized_responses.append(response.model_copy(update={"proposals": normalized_proposals}))
        detail_responses = normalized_responses
    finally:
        value_book.close()
        formula_book.close()
    validations = [
        validate_proposal(proposal, total_rows=total_rows, total_columns=total_columns)
        for response in detail_responses
        for proposal in response.proposals
    ]
    for validation in validations:
        notes = normalization_notes.get(validation.proposal_id, [])
        if notes:
            validation.outcome = ValidationOutcome.ADJUSTED
            validation.corrections.extend(notes)
    return outline_responses, detail_responses, validations
