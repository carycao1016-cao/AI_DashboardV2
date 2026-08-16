"""根据已通过物理校验的边界提取源单元格，不让 AI 生成业务数值。"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

from parser_poc.contracts import BoundaryValidationResult, SignificanceLayout, TableBoundaryProposal, ValidationOutcome
from parser_poc.workbook_scan import approximate_display_value, cell_type, json_safe


_MISSING_TEXT = {"-", "- ", "—", "–"}
_HEADER_CODE_PATTERN = re.compile(r"\(([A-Za-z]+)\)\s*$")
_STANDALONE_CODE_PATTERN = re.compile(r"^[A-Za-z]+$")


def _availability(value: Any) -> str:
    return "not_available" if value is None or (isinstance(value, str) and value.strip() in _MISSING_TEXT) else "available"


def _parse_value(value: Any, number_format: str) -> tuple[Any, str, str]:
    """只在 Excel 格式或源文本提供充分证据时解析单位。"""
    if _availability(value) == "not_available":
        return None, "", "unknown"
    if isinstance(value, (int, float)):
        if "%" in (number_format or ""):
            return value, "percentage", "excel_stored_value"
        return value, "", "excel_stored_value"
    text = str(value).strip()
    if text.endswith("%"):
        try:
            return float(text[:-1]) / 100, "percentage", "text_parsed"
        except ValueError:
            pass
    return text, "", "text_parsed"


def _question_text(sheet: Any, title_rows: list[int]) -> str:
    for row in title_rows:
        value = sheet.cell(row, 1).value
        if value not in (None, "") and str(value).strip() != "#page":
            return str(value).strip()
    return ""


def _question_number(text: str) -> str:
    match = re.match(r"\s*([A-Za-z]?\d+[A-Za-z]?)", text)
    return match.group(1) if match else ""


def _header_significance_code(path_values: list[Any]) -> str:
    """从独立代码行或 ``Total (A)`` 形式的表头提取比较代码。"""
    for value in reversed(path_values):
        if value in (None, ""):
            continue
        text = str(value).strip()
        parenthesized = _HEADER_CODE_PATTERN.search(text)
        if parenthesized:
            return parenthesized.group(1)
        if _STANDALONE_CODE_PATTERN.fullmatch(text):
            return text
    return ""


def _data_columns_for_layout(
    *,
    sheet: Any,
    min_column: int,
    max_column: int,
    header_rows: list[int],
    layout: SignificanceLayout,
) -> list[int]:
    """确定真正承载数值的列，避免把 Decipher 相邻显著性列作为指标列。"""
    columns = list(range(max(2, min_column), max_column + 1))
    if layout != SignificanceLayout.ADJACENT_COLUMN:
        return columns
    data_columns = []
    for column in columns:
        # Decipher 的 marker/间隔列没有自己的表头；合并单元格在 read_only
        # 模式下仅左上角有值，因此不能用“上一列有父表头”来推断其为数值列。
        has_own_header = any(sheet.cell(row, column).value not in (None, "") for row in header_rows)
        if has_own_header:
            data_columns.append(column)
    return data_columns


def extract_validated_table(
    *,
    path: Path,
    proposal: TableBoundaryProposal,
    validation: BoundaryValidationResult,
    extracted_table_id: str,
    metric_type: str = "unknown",
) -> dict[str, Any]:
    """回读一张表；非 accepted 结果只能生成 review_required 提取包。"""
    if not extracted_table_id:
        raise ValueError("extracted_table_id is required")
    workbook_values = load_workbook(path, read_only=True, data_only=True)
    workbook_formulas = load_workbook(path, read_only=True, data_only=False)
    try:
        if proposal.sheet_name not in workbook_values.sheetnames:
            raise ValueError("Sheet not found: %s" % proposal.sheet_name)
        value_sheet = workbook_values[proposal.sheet_name]
        formula_sheet = workbook_formulas[proposal.sheet_name]
        min_col, min_row, max_col, max_row = range_boundaries(proposal.source_range)
        if validation.outcome not in {ValidationOutcome.ACCEPTED, ValidationOutcome.ADJUSTED}:
            review_status = "review_required"
        else:
            review_status = "auto_approved"
        data_columns = _data_columns_for_layout(
            sheet=value_sheet,
            min_column=min_col,
            max_column=max_col,
            header_rows=proposal.regions.header_rows,
            layout=proposal.regions.significance_layout,
        )
        header_ids: dict[int, str] = {}
        header_items = []
        label_map: dict[str, str] = {}
        last_header_row = proposal.regions.header_rows[-1] if proposal.regions.header_rows else None
        for column in data_columns:
            path_values = [value_sheet.cell(row, column).value for row in proposal.regions.header_rows]
            non_empty_path = [str(value).strip() for value in path_values if value not in (None, "")]
            code = _header_significance_code(path_values)
            header_id = "%s_H%s" % (extracted_table_id, column)
            header_ids[column] = header_id
            if code:
                label_map[code] = header_id
            header_items.append(
                {
                    "extracted_header_id": header_id,
                    "data_column": get_column_letter(column),
                    "header_path": non_empty_path,
                    "display_label": non_empty_path[0] if non_empty_path else "",
                    "significance_code": code,
                    "source_header_cells": ["%s%s" % (get_column_letter(column), row) for row in proposal.regions.header_rows],
                }
            )

        significance_rows = set(proposal.regions.significance_locations)
        rows = []
        all_value_rows = [*proposal.regions.base_rows, *proposal.regions.data_rows]
        for row_number in all_value_rows:
            if row_number < min_row or row_number > max_row:
                continue
            row_kind = "base" if row_number in proposal.regions.base_rows else "data"
            row_id = "%s_R%s" % (extracted_table_id, row_number)
            cells = []
            for column in data_columns:
                value_cell = value_sheet.cell(row_number, column)
                formula_cell = formula_sheet.cell(row_number, column)
                raw_value = json_safe(value_cell.value)
                display_value, display_source = approximate_display_value(value_cell.value, value_cell.number_format)
                parsed_value, parsed_unit, precision_source = _parse_value(value_cell.value, value_cell.number_format)
                marker = ""
                marker_source = ""
                # Quantum 等表格把显著性代码放在数据行的下一行；模型可能将这种布局
                # 命名为 following_row，也可能命名为 separate_label_row。两者在源表
                # 读取层的物理含义相同，均必须按 significance_locations 严格取值。
                if row_kind == "data" and proposal.regions.significance_layout in {
                    SignificanceLayout.FOLLOWING_ROW,
                    SignificanceLayout.SEPARATE_LABEL_ROW,
                }:
                    marker_row = row_number + 1
                    if marker_row in significance_rows:
                        marker_value = value_sheet.cell(marker_row, column).value
                        marker = "" if marker_value in (None, "") else str(marker_value).strip()
                        marker_source = "%s%s" % (get_column_letter(column), marker_row) if marker else ""
                elif row_kind == "data" and proposal.regions.significance_layout == SignificanceLayout.ADJACENT_COLUMN:
                    marker_column = column + 1
                    if marker_column <= max_col:
                        marker_value = value_sheet.cell(row_number, marker_column).value
                        marker = "" if marker_value in (None, "") else str(marker_value).strip()
                        marker_source = "%s%s" % (get_column_letter(marker_column), row_number) if marker else ""
                referenced_ids = [label_map[code] for code in marker if code in label_map]
                cells.append(
                    {
                        "extracted_cell_id": "%s_C%s" % (row_id, column),
                        "extracted_row_id": row_id,
                        "extracted_header_id": header_ids[column],
                        "source_cell": value_cell.coordinate,
                        "raw_value": raw_value,
                        "raw_type": cell_type(value_cell.value, value_cell.data_type),
                        "excel_display_value": display_value,
                        "excel_display_value_source": display_source,
                        "excel_number_format": value_cell.number_format,
                        "formula": formula_cell.value if formula_cell.data_type == "f" else None,
                        "parsed_value": parsed_value,
                        "parsed_unit": parsed_unit or metric_type if parsed_value is not None else "",
                        "precision_source": precision_source,
                        "availability_status": _availability(value_cell.value),
                        "original_significance_marker": marker,
                        "significance_marker_source_cell": marker_source,
                        "significance_representation": proposal.regions.significance_layout.value,
                        "significance_referenced_header_ids": referenced_ids,
                        "significance_mapping_status": "mapped" if marker and len(referenced_ids) == len(marker) else ("unresolved" if marker else "not_applicable"),
                        "is_suppressed": _availability(value_cell.value) == "not_available",
                        "review_status": review_status,
                    }
                )
            rows.append(
                {
                    "extracted_row_id": row_id,
                    "row_index": row_number,
                    "source_range": "%s%s:%s%s" % (get_column_letter(min_col), row_number, get_column_letter(max_col), row_number),
                    "original_label": json_safe(value_sheet.cell(row_number, 1).value),
                    "detected_row_type": row_kind,
                    "detected_metric_type": metric_type,
                    "cells": cells,
                    "review_status": review_status,
                }
            )
        title_text = _question_text(value_sheet, proposal.regions.title_rows)
        return {
            "schema_name": "extracted_table",
            "schema_version": "0.1.0-poc",
            "extracted_table_id": extracted_table_id,
            "source_sheet": proposal.sheet_name,
            "source_range": proposal.source_range,
            "detected_question_number": _question_number(title_text),
            "detected_question_text": title_text,
            "detected_table_title": title_text,
            "table_variant": metric_type,
            "structure": {
                "title_rows": proposal.regions.title_rows,
                "header_rows": proposal.regions.header_rows,
                "base_rows": proposal.regions.base_rows,
                "data_rows": proposal.regions.data_rows,
                "footnote_rows": proposal.regions.footnote_rows,
                "data_column_start": min(data_columns) if data_columns else None,
                "data_column_end": max(data_columns) if data_columns else None,
            },
            "headers": header_items,
            "rows": rows,
            "significance_schema": {
                "presence": "present" if proposal.regions.significance_layout != SignificanceLayout.NONE else "not_available",
                "layout": proposal.regions.significance_layout.value,
                "label_map": label_map,
                "parse_status": "mapped" if label_map else "not_applicable",
            },
            "validation_outcome": validation.outcome.value,
            "review_status": review_status,
        }
    finally:
        workbook_values.close()
        workbook_formulas.close()
