"""Generate a neutral, bounded WorkbookScanSummary for the table-boundary PoC."""

from __future__ import annotations

import argparse
import csv
import json
import math
import re
import zipfile
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, range_boundaries


DEFAULT_TARGET_TOKENS = 64_000
DEFAULT_HARD_TOKENS = 70_000
MAX_SAMPLES_PER_ROW = 6
MAX_CONTEXT_REQUEST_ROWS = 100
MAX_CONTINUATION_REQUESTS = 2

OUTLINE_ROW_SCHEMA = [
    "row_number",
    "a_value",
    "first_non_empty_cell",
    "first_non_empty_value",
    "non_empty_count",
    "non_empty_column_range",
    "text_count",
    "numeric_count",
    "percentage_like_count",
]

ENCODING_CANDIDATES = ("utf-8-sig", "utf-8", "gb18030", "gbk", "big5")


def json_safe(value: Any) -> Any:
    """Convert openpyxl values to JSON-safe, source-preserving primitives."""
    if value is None or isinstance(value, (str, int, bool)):
        return value
    if isinstance(value, float):
        return value if math.isfinite(value) else str(value)
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def column_range(columns: Sequence[int]) -> Optional[str]:
    if not columns:
        return None
    return "%s:%s" % (get_column_letter(min(columns)), get_column_letter(max(columns)))


def approximate_display_value(value: Any, number_format: str) -> Tuple[Optional[str], str]:
    """Return a safe best-effort display value; never claim exact Excel rendering."""
    if value is None:
        return None, "empty"
    if isinstance(value, str):
        return value, "source_text"
    if isinstance(value, (int, float)) and "%" in (number_format or ""):
        return "%s%%" % ("%g" % (value * 100)), "number_format_approximation"
    return str(value), "raw_value_string"


def cell_type(value: Any, data_type: str) -> str:
    if data_type == "e":
        return "error"
    if value is None:
        return "empty"
    if isinstance(value, bool):
        return "boolean"
    if isinstance(value, (int, float)):
        return "number"
    return "text"


def is_percentage_like(value: Any, number_format: str) -> bool:
    return isinstance(value, (int, float)) and "%" in (number_format or "")


def select_sample_columns(non_empty_columns: Sequence[int], a_is_non_empty: bool) -> List[int]:
    """Choose up to six cells by position, without inspecting their business meaning."""
    columns = sorted(set(non_empty_columns))
    if not columns:
        return []

    selected: List[int] = [1] if a_is_non_empty else []
    positions = [0, 1, (len(columns) - 1) // 2, len(columns) // 2, len(columns) - 2, len(columns) - 1]
    for position in positions:
        if 0 <= position < len(columns) and columns[position] not in selected:
            selected.append(columns[position])
        if len(selected) >= MAX_SAMPLES_PER_ROW:
            break
    return sorted(selected[:MAX_SAMPLES_PER_ROW])


def attribute(cell: Any, name: str, default: Any = None) -> Any:
    return getattr(cell, name, default)


def row_summary(value_cells: Sequence[Any], formula_cells: Sequence[Any], row_number: int) -> Dict[str, Any]:
    cells = []
    non_empty_columns = []
    text_count = numeric_count = percentage_like_count = 0

    for column, value_cell in enumerate(value_cells, 1):
        value = attribute(value_cell, "value")
        if value is None:
            continue
        non_empty_columns.append(column)
        number_format = attribute(value_cell, "number_format", "General")
        kind = cell_type(value, attribute(value_cell, "data_type", "s"))
        if kind == "text":
            text_count += 1
        elif kind == "number":
            numeric_count += 1
            if is_percentage_like(value, number_format):
                percentage_like_count += 1

    if not non_empty_columns:
        return {"kind": "blank_row", "row_number": row_number}

    for column in select_sample_columns(non_empty_columns, 1 in non_empty_columns):
        value_cell = value_cells[column - 1]
        formula_cell = formula_cells[column - 1]
        value = attribute(value_cell, "value")
        number_format = attribute(value_cell, "number_format", "General")
        display_value, display_source = approximate_display_value(value, number_format)
        has_formula = attribute(formula_cell, "data_type") == "f"
        cells.append(
            {
                "cell": "%s%s" % (get_column_letter(column), row_number),
                "raw_value": json_safe(value),
                "display_value": display_value,
                "display_value_source": display_source,
                "data_type": cell_type(value, attribute(value_cell, "data_type", "s")),
                "number_format": number_format,
                "has_formula": has_formula,
                "formula_result_available": bool(has_formula and value is not None),
                "error_code": str(value) if attribute(value_cell, "data_type") == "e" else None,
            }
        )

    first_column = min(non_empty_columns)
    first_cell = value_cells[first_column - 1]
    a_value = json_safe(attribute(value_cells[0], "value")) if value_cells else None
    return {
        "kind": "row",
        "row_number": row_number,
        "a_value": a_value,
        "first_non_empty": {
            "cell": None if first_column == 1 else "%s%s" % (get_column_letter(first_column), row_number),
            "value": None if first_column == 1 else json_safe(attribute(first_cell, "value")),
        },
        "non_empty_count": len(non_empty_columns),
        "non_empty_column_range": column_range(non_empty_columns),
        "text_count": text_count,
        "numeric_count": numeric_count,
        "percentage_like_count": percentage_like_count,
        "sampled_cells": cells,
    }


def outline_row_summary(value_cells: Sequence[Any], row_number: int) -> Dict[str, Any]:
    """Build the first-pass outline without exposing per-cell samples to AI."""
    non_empty_columns = []
    text_count = numeric_count = percentage_like_count = 0
    for column, value_cell in enumerate(value_cells, 1):
        value = attribute(value_cell, "value")
        if value is None:
            continue
        non_empty_columns.append(column)
        number_format = attribute(value_cell, "number_format", "General")
        kind = cell_type(value, attribute(value_cell, "data_type", "s"))
        if kind == "text":
            text_count += 1
        elif kind == "number":
            numeric_count += 1
            if is_percentage_like(value, number_format):
                percentage_like_count += 1
    if not non_empty_columns:
        return {"kind": "blank_row", "row_number": row_number}
    first_column = min(non_empty_columns)
    first_cell = value_cells[first_column - 1]
    a_value = json_safe(attribute(value_cells[0], "value")) if value_cells else None
    return {
        "kind": "row",
        "row_number": row_number,
        "a_value": a_value,
        "first_non_empty": {
            "cell": None if first_column == 1 else "%s%s" % (get_column_letter(first_column), row_number),
            "value": None if first_column == 1 else json_safe(attribute(first_cell, "value")),
        },
        "non_empty_count": len(non_empty_columns),
        "non_empty_column_range": column_range(non_empty_columns),
        "text_count": text_count,
        "numeric_count": numeric_count,
        "percentage_like_count": percentage_like_count,
    }


def compact_blank_rows(rows: Sequence[Dict[str, Any]]) -> List[Dict[str, Any]]:
    compacted: List[Dict[str, Any]] = []
    index = 0
    while index < len(rows):
        current = rows[index]
        if current["kind"] != "blank_row":
            compacted.append(current)
            index += 1
            continue
        start = current["row_number"]
        end = start
        while index + 1 < len(rows) and rows[index + 1]["kind"] == "blank_row":
            index += 1
            end = rows[index]["row_number"]
        if start == end:
            compacted.append(current)
        else:
            compacted.append(
                {
                    "kind": "blank_row_range",
                    "start_row": start,
                    "end_row": end,
                    "row_count": end - start + 1,
                }
            )
        index += 1
    return compacted


def estimate_tokens(payload: Any) -> int:
    serialized = json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
    return math.ceil(len(serialized) / 3.5)


def relevant_merged_ranges(merged_ranges: Sequence[Dict[str, Any]], start_row: int, end_row: int) -> List[Dict[str, Any]]:
    return [
        item
        for item in merged_ranges
        if item["min_row"] <= end_row and item["max_row"] >= start_row
    ]


def xlsx_sheet_metadata(path: Path) -> Dict[str, Dict[str, Any]]:
    """Read layout metadata directly from XLSX XML without materializing cells."""
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    package_rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    namespaces = {"main": main_ns, "rel": rel_ns, "pkg": package_rel_ns}
    metadata: Dict[str, Dict[str, Any]] = {}

    with zipfile.ZipFile(path) as archive:
        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        relationships = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        targets = {
            relation.attrib["Id"]: relation.attrib["Target"].lstrip("/")
            for relation in relationships.findall("pkg:Relationship", namespaces)
        }
        for sheet in workbook.findall("main:sheets/main:sheet", namespaces):
            name = sheet.attrib["name"]
            relationship_id = sheet.attrib["{%s}id" % rel_ns]
            target = targets[relationship_id]
            xml_path = target if target.startswith("xl/") else "xl/" + target
            tree = ET.fromstring(archive.read(xml_path))
            merges = []
            for merge in tree.findall("main:mergeCells/main:mergeCell", namespaces):
                reference = merge.attrib["ref"]
                min_col, min_row, _max_col, max_row = range_boundaries(reference)
                merges.append(
                    {
                        "range": reference,
                        "top_left_cell": "%s%s" % (get_column_letter(min_col), min_row),
                        "top_left_display_value": None,
                        "display_value_source": "not_loaded_from_layout_xml",
                        "min_row": min_row,
                        "max_row": max_row,
                    }
                )
            hidden_row_numbers = [
                int(row.attrib["r"])
                for row in tree.findall("main:sheetData/main:row", namespaces)
                if row.attrib.get("hidden") in {"1", "true"}
            ]
            hidden_column_letters = []
            for column in tree.findall("main:cols/main:col", namespaces):
                if column.attrib.get("hidden") in {"1", "true"}:
                    hidden_column_letters.extend(
                        get_column_letter(index)
                        for index in range(int(column.attrib["min"]), int(column.attrib["max"]) + 1)
                    )
            metadata[name] = {
                "is_hidden_sheet": sheet.attrib.get("state") in {"hidden", "veryHidden"},
                "merged_ranges": merges,
                "hidden_rows": hidden_row_numbers,
                "hidden_columns": hidden_column_letters,
            }
    return metadata


def build_chunks_from_rows(
    rows: Sequence[Dict[str, Any]],
    sheet_name: str,
    total_sheet_rows: int,
    total_sheet_columns: int,
    merged_ranges: Sequence[Dict[str, Any]],
    hidden_row_numbers: Sequence[int],
    hidden_column_letters: Sequence[str],
    target_tokens: int,
    hard_tokens: int,
    compact_outline: bool = False,
) -> List[Dict[str, Any]]:
    chunks: List[Dict[str, Any]] = []
    pending: List[Dict[str, Any]] = []
    start_row: Optional[int] = None

    def materialize(items: Sequence[Dict[str, Any]], start: int, end: int) -> Dict[str, Any]:
        payload = {
            "chunk_row_start": start,
            "chunk_row_end": end,
            "total_sheet_rows": total_sheet_rows,
            "total_sheet_columns": total_sheet_columns,
            "merged_ranges": relevant_merged_ranges(merged_ranges, start, end),
            "hidden_rows": [row for row in hidden_row_numbers if start <= row <= end],
            "hidden_columns": list(hidden_column_letters),
        }
        if compact_outline:
            payload["row_schema"] = OUTLINE_ROW_SCHEMA
            payload["rows"] = [
                [
                    item["row_number"],
                    item["a_value"],
                    item["first_non_empty"]["cell"],
                    item["first_non_empty"]["value"],
                    item["non_empty_count"],
                    item["non_empty_column_range"],
                    item["text_count"],
                    item["numeric_count"],
                    item["percentage_like_count"],
                ]
                for item in items
                if item["kind"] == "row"
            ]
            payload["blank_rows"] = [
                [item["row_number"], item["row_number"]]
                if item["kind"] == "blank_row"
                else [item["start_row"], item["end_row"]]
                for item in items
                if item["kind"] != "row"
            ]
        else:
            payload["rows"] = list(items)
        return payload

    for item in rows:
        item_start = item.get("row_number", item.get("start_row"))
        item_end = item.get("row_number", item.get("end_row"))
        if start_row is None:
            start_row = item_start
        candidate = pending + [item]
        candidate_payload = materialize(candidate, start_row, item_end)
        if pending and estimate_tokens(candidate_payload) > target_tokens:
            previous_end = pending[-1].get("row_number", pending[-1].get("end_row"))
            chunks.append(materialize(pending, start_row, previous_end))
            pending = [item]
            start_row = item_start
            if estimate_tokens(materialize(pending, start_row, item_end)) > hard_tokens:
                raise ValueError("A single summary item exceeds the hard token budget at row %s" % item_start)
        else:
            pending = candidate

    if pending and start_row is not None:
        end_row = pending[-1].get("row_number", pending[-1].get("end_row"))
        chunks.append(materialize(pending, start_row, end_row))

    for index, chunk in enumerate(chunks, 1):
        chunk["chunk_id"] = "%s_chunk_%03d_rows_%s_%s" % (
            sheet_name,
            index,
            chunk["chunk_row_start"],
            chunk["chunk_row_end"],
        )
        chunk["estimated_input_tokens"] = estimate_tokens(chunk)
        chunk["ai_context_policy"] = {
            "max_context_rows_per_side": MAX_CONTEXT_REQUEST_ROWS,
            "max_continuation_requests": MAX_CONTINUATION_REQUESTS,
            "table_boundary_not_inferred_by_python": True,
        }
    return chunks


def build_sheet_chunks(
    value_sheet: Any,
    formula_sheet: Any,
    layout_metadata: Dict[str, Any],
    target_tokens: int,
    hard_tokens: int,
) -> List[Dict[str, Any]]:
    summaries = []
    for row_number, (value_cells, formula_cells) in enumerate(zip(value_sheet.iter_rows(), formula_sheet.iter_rows()), 1):
        summaries.append(row_summary(value_cells, formula_cells, row_number))
    return build_chunks_from_rows(
        compact_blank_rows(summaries),
        value_sheet.title,
        value_sheet.max_row,
        value_sheet.max_column,
        layout_metadata["merged_ranges"],
        layout_metadata["hidden_rows"],
        layout_metadata["hidden_columns"],
        target_tokens,
        hard_tokens,
    )


def build_sheet_outline_chunks(
    value_sheet: Any,
    layout_metadata: Dict[str, Any],
    target_tokens: int,
    hard_tokens: int,
) -> List[Dict[str, Any]]:
    summaries = [
        outline_row_summary(value_cells, row_number)
        for row_number, value_cells in enumerate(value_sheet.iter_rows(), 1)
    ]
    return build_chunks_from_rows(
        compact_blank_rows(summaries),
        value_sheet.title,
        value_sheet.max_row,
        value_sheet.max_column,
        layout_metadata["merged_ranges"],
        layout_metadata["hidden_rows"],
        layout_metadata["hidden_columns"],
        target_tokens,
        hard_tokens,
        compact_outline=True,
    )


def parse_row_range(value: str) -> Tuple[int, int]:
    match = re.fullmatch(r"(\d+):(\d+)", value.strip())
    if not match:
        raise ValueError("Detail ranges must use start:end rows, for example 120:185")
    start_row, end_row = (int(match.group(1)), int(match.group(2)))
    if start_row > end_row:
        raise ValueError("Detail range start must not exceed end: %s" % value)
    return start_row, end_row


def plan_detail_windows(
    candidate_ranges: Sequence[Tuple[int, int]],
    total_rows: int,
    context_before: int,
    context_after: int,
    max_candidate_gap_rows: int,
) -> List[Dict[str, Any]]:
    """Group nearby AI-proposed ranges by position, without deciding table identity."""
    if context_before < 0 or context_after < 0 or max_candidate_gap_rows < 0:
        raise ValueError("Detail window configuration values must be non-negative")
    candidates = sorted(candidate_ranges)
    windows: List[List[Tuple[int, int]]] = []
    for candidate in candidates:
        if candidate[0] < 1 or candidate[1] > total_rows:
            raise ValueError("Candidate range is outside the Sheet: %s:%s" % candidate)
        if not windows or candidate[0] - windows[-1][-1][1] - 1 > max_candidate_gap_rows:
            windows.append([candidate])
        else:
            windows[-1].append(candidate)

    return [
        {
            "window_id": "detail_window_%03d" % index,
            "candidate_ranges": ["%s:%s" % candidate for candidate in candidates_in_window],
            "window_row_start": max(1, candidates_in_window[0][0] - context_before),
            "window_row_end": min(total_rows, candidates_in_window[-1][1] + context_after),
            "context_before_rows": context_before,
            "context_after_rows": context_after,
            "max_candidate_gap_rows": max_candidate_gap_rows,
        }
        for index, candidates_in_window in enumerate(windows, 1)
    ]


def build_detail_window(
    value_sheet: Any,
    formula_sheet: Any,
    layout_metadata: Dict[str, Any],
    window: Dict[str, Any],
    target_tokens: int,
    hard_tokens: int,
) -> Dict[str, Any]:
    start_row, end_row = window["window_row_start"], window["window_row_end"]
    value_rows = value_sheet.iter_rows(min_row=start_row, max_row=end_row)
    formula_rows = formula_sheet.iter_rows(min_row=start_row, max_row=end_row)
    summaries = [
        row_summary(value_cells, formula_cells, row_number)
        for row_number, (value_cells, formula_cells) in enumerate(zip(value_rows, formula_rows), start_row)
    ]
    detail = dict(window)
    detail["detail_chunks"] = build_chunks_from_rows(
        compact_blank_rows(summaries),
        value_sheet.title,
        value_sheet.max_row,
        value_sheet.max_column,
        layout_metadata["merged_ranges"],
        layout_metadata["hidden_rows"],
        layout_metadata["hidden_columns"],
        target_tokens,
        hard_tokens,
    )
    return detail


def scan_xlsx(
    path: Path,
    target_tokens: int,
    hard_tokens: int,
    sheet_name: Optional[str],
    detail_ranges: Sequence[Tuple[int, int]],
    detail_context_before: int,
    detail_context_after: int,
    max_candidate_gap_rows: int,
) -> Dict[str, Any]:
    layout = xlsx_sheet_metadata(path)
    value_book = load_workbook(path, read_only=True, data_only=True)
    requested_sheets = [sheet_name] if sheet_name else value_book.sheetnames
    if detail_ranges and not sheet_name:
        raise ValueError("--detail-range requires --sheet")
    sheets = []
    for name in requested_sheets:
        if name not in value_book.sheetnames:
            raise ValueError("Sheet not found: %s" % name)
        value_sheet = value_book[name]
        sheet_payload = {
            "sheet_name": name,
            "used_range": "A1:%s%s" % (get_column_letter(value_sheet.max_column), value_sheet.max_row),
            "is_hidden_sheet": layout[name]["is_hidden_sheet"],
            "outline_chunks": build_sheet_outline_chunks(value_sheet, layout[name], target_tokens, hard_tokens),
        }
        if detail_ranges:
            formula_book = load_workbook(path, read_only=True, data_only=False)
            windows = plan_detail_windows(
                detail_ranges,
                value_sheet.max_row,
                detail_context_before,
                detail_context_after,
                max_candidate_gap_rows,
            )
            sheet_payload["detail_windows"] = [
                build_detail_window(value_sheet, formula_book[name], layout[name], window, target_tokens, hard_tokens)
                for window in windows
            ]
        sheets.append(sheet_payload)
    return {
        "schema_name": "workbook_scan_summary",
        "schema_version": "0.1.0-poc",
        "workbook_name_hint": path.name,
        "workbook_name_is_evidence": False,
        "source_format": "xlsx",
        "sheets": sheets,
    }


def decode_csv(path: Path) -> Tuple[str, str, float]:
    raw = path.read_bytes()
    successful: List[Tuple[float, str, str]] = []
    for encoding in ENCODING_CANDIDATES:
        try:
            decoded = raw.decode(encoding)
        except UnicodeDecodeError:
            continue
        replacement_penalty = decoded.count("\ufffd") * 10
        control_penalty = sum(ord(char) < 32 and char not in "\n\r\t" for char in decoded)
        cjk_or_latin = sum(char.isascii() and char.isalpha() or "\u4e00" <= char <= "\u9fff" for char in decoded)
        score = cjk_or_latin - replacement_penalty - control_penalty
        successful.append((score, encoding, decoded))
    if not successful:
        raise ValueError("No approved encoding decoded the CSV")
    score, encoding, decoded = max(successful, key=lambda item: item[0])
    confidence = 1.0 if len(successful) == 1 else 0.7
    return decoded, encoding, confidence


def scan_csv(path: Path, target_tokens: int, hard_tokens: int) -> Dict[str, Any]:
    decoded, encoding, confidence = decode_csv(path)
    rows = list(csv.reader(decoded.splitlines()))

    class CsvCell:
        def __init__(self, value: Optional[str]) -> None:
            self.value = value if value != "" else None
            self.data_type = "s"
            self.number_format = "General"

    max_column = max((len(row) for row in rows), default=1)
    summaries = []
    for row_number, values in enumerate(rows, 1):
        cells = [CsvCell(values[index] if index < len(values) else None) for index in range(max_column)]
        summaries.append(outline_row_summary(cells, row_number))

    return {
        "schema_name": "workbook_scan_summary",
        "schema_version": "0.1.0-poc",
        "workbook_name_hint": path.name,
        "workbook_name_is_evidence": False,
        "source_format": "csv",
        "csv_encoding": {"selected": encoding, "confidence": confidence},
        "sheets": [
            {
                "sheet_name": path.stem,
                "used_range": "A1:%s%s" % (get_column_letter(max_column), len(rows)),
                "is_hidden_sheet": False,
                "outline_chunks": build_chunks_from_rows(
                    compact_blank_rows(summaries),
                    path.stem,
                    len(rows),
                    max_column,
                    [],
                    [],
                    [],
                    target_tokens,
                    hard_tokens,
                    compact_outline=True,
                ),
            }
        ],
    }


def scan_workbook(
    path: Path,
    target_tokens: int = DEFAULT_TARGET_TOKENS,
    hard_tokens: int = DEFAULT_HARD_TOKENS,
    sheet_name: Optional[str] = None,
    detail_ranges: Sequence[Tuple[int, int]] = (),
    detail_context_before: int = 20,
    detail_context_after: int = 20,
    max_candidate_gap_rows: int = 20,
) -> Dict[str, Any]:
    if hard_tokens < target_tokens:
        raise ValueError("hard_tokens must be greater than or equal to target_tokens")
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        return scan_xlsx(
            path,
            target_tokens,
            hard_tokens,
            sheet_name,
            detail_ranges,
            detail_context_before,
            detail_context_after,
            max_candidate_gap_rows,
        )
    if suffix == ".csv":
        if sheet_name:
            raise ValueError("--sheet is not supported for CSV")
        return scan_csv(path, target_tokens, hard_tokens)
    raise ValueError("Unsupported file type: %s" % suffix)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("input", type=Path)
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--sheet")
    parser.add_argument("--target-tokens", type=int, default=DEFAULT_TARGET_TOKENS)
    parser.add_argument("--hard-tokens", type=int, default=DEFAULT_HARD_TOKENS)
    parser.add_argument("--detail-range", action="append", default=[], help="AI-proposed candidate rows as start:end; repeatable")
    parser.add_argument("--detail-context-before", type=int, default=20)
    parser.add_argument("--detail-context-after", type=int, default=20)
    parser.add_argument("--max-candidate-gap-rows", type=int, default=20)
    args = parser.parse_args()
    summary = scan_workbook(
        args.input,
        args.target_tokens,
        args.hard_tokens,
        args.sheet,
        [parse_row_range(item) for item in args.detail_range],
        args.detail_context_before,
        args.detail_context_after,
        args.max_candidate_gap_rows,
    )
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print("Wrote %s" % args.output)


if __name__ == "__main__":
    main()
