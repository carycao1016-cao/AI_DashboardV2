import unittest
from pathlib import Path

from parser_poc.contracts import (
    CoarseRange,
    DetailWindowResponse,
    OutlineStatus,
    SheetOutlineResponse,
    SignificanceLayout,
    TableBoundaryProposal,
    TableRegions,
    ValidationOutcome,
)
from parser_poc.orchestration import ControlledStubAdapter, normalize_edge_rows, normalize_known_summary_rows, run_sheet_with_adapter, run_xlsx_sheet_with_adapter, validate_proposal
from openpyxl import load_workbook


class OrchestrationTests(unittest.TestCase):
    def test_controlled_stub_runs_two_layers_and_validates_physical_range(self):
        outline = SheetOutlineResponse(
            sheet_name="Sheet1",
            chunk_id="Sheet1_chunk_001_rows_1_50",
            candidates=[CoarseRange(candidate_id="candidate_1", start_row=5, end_row=25, status=OutlineStatus.COMPLETE)],
        )
        detail = DetailWindowResponse(
            sheet_name="Sheet1",
            window_id="detail_window_001",
            proposals=[
                TableBoundaryProposal(
                    proposal_id="proposal_1",
                    candidate_id="candidate_1",
                    sheet_name="Sheet1",
                    source_range="A5:F25",
                    regions=TableRegions(header_rows=[5, 6], data_rows=list(range(7, 25))),
                )
            ],
        )
        adapter = ControlledStubAdapter([outline, detail])
        outlines, details, validations = run_sheet_with_adapter(
            sheet_name="Sheet1",
            total_rows=50,
            total_columns=6,
            outline_chunks=[{"chunk_id": "Sheet1_chunk_001_rows_1_50", "rows": []}],
            adapter=adapter,
        )
        self.assertEqual(len(outlines), 1)
        self.assertEqual(len(details), 1)
        self.assertEqual(validations[0].outcome, ValidationOutcome.ACCEPTED)
        self.assertEqual([call["task_name"] for call in adapter.calls], ["sheet_outline", "detail_window"])

    def test_out_of_bounds_proposal_requires_review(self):
        proposal = TableBoundaryProposal(
            proposal_id="proposal_2",
            candidate_id="candidate_2",
            sheet_name="Sheet1",
            source_range="A5:F55",
            regions=TableRegions(header_rows=[5], data_rows=[6]),
        )
        result = validate_proposal(proposal, total_rows=50, total_columns=6)
        self.assertEqual(result.outcome, ValidationOutcome.REVIEW_REQUIRED)

    def test_significance_overlap_requires_declared_layout(self):
        with self.assertRaises(ValueError):
            TableRegions(header_rows=[5], data_rows=[6], significance_locations=[6])
        regions = TableRegions(
            header_rows=[5],
            data_rows=[6],
            significance_locations=[6],
            significance_layout=SignificanceLayout.FOLLOWING_ROW,
        )
        self.assertEqual(regions.significance_layout, SignificanceLayout.FOLLOWING_ROW)

    def test_real_xlsx_runner_sends_detail_samples_only_after_outline_candidate(self):
        outline = SheetOutlineResponse(
            sheet_name="ban1_%Sig",
            chunk_id="ban1_%Sig_chunk_001_rows_1_78",
            candidates=[CoarseRange(candidate_id="candidate_1", start_row=1, end_row=78, status=OutlineStatus.COMPLETE)],
        )
        detail = DetailWindowResponse(
            sheet_name="ban1_%Sig",
            window_id="detail_window_001",
            proposals=[
                TableBoundaryProposal(
                    proposal_id="proposal_1",
                    candidate_id="candidate_1",
                    sheet_name="ban1_%Sig",
                    source_range="A2:AB77",
                    regions=TableRegions(
                        header_rows=[10, 11],
                        base_rows=[13],
                        data_rows=list(range(14, 75)),
                        footnote_rows=[77],
                        significance_locations=[15],
                        significance_layout=SignificanceLayout.FOLLOWING_ROW,
                    ),
                )
            ],
        )
        adapter = ControlledStubAdapter([outline, detail])
        path = Path(__file__).parents[2] / "outputs/ark_smoke/Tabs_%95_first_table.xlsx"
        outlines, details, validations = run_xlsx_sheet_with_adapter(
            path=path,
            sheet_name="ban1_%Sig",
            adapter=adapter,
        )
        self.assertEqual(len(outlines), 1)
        self.assertEqual(len(details), 1)
        self.assertEqual(validations[0].outcome, ValidationOutcome.ACCEPTED)
        self.assertIn("detail_chunks", adapter.calls[1]["payload"])
        self.assertTrue(adapter.calls[1]["payload"]["detail_chunks"][0]["rows"])

    def test_normalize_edge_rows_removes_page_marker_and_trailing_blank(self):
        proposal = TableBoundaryProposal(
            proposal_id="proposal_edges",
            candidate_id="candidate_edges",
            sheet_name="ban1_%Sig",
            source_range="A1:AB78",
            regions=TableRegions(
                title_rows=[2, 3, 4, 5],
                header_rows=[8, 10, 11],
                base_rows=[13],
                data_rows=[14, 16],
                footnote_rows=[77],
                significance_locations=[15, 17],
                significance_layout=SignificanceLayout.SEPARATE_LABEL_ROW,
            ),
        )
        path = Path(__file__).parents[2] / "outputs/ark_smoke/Tabs_%95_first_table.xlsx"
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            normalized, corrections = normalize_edge_rows(proposal, workbook["ban1_%Sig"])
        finally:
            workbook.close()
        self.assertEqual(normalized.source_range, "A2:AB77")
        self.assertEqual(corrections, ["trimmed_top_page_marker_row_1", "trimmed_bottom_blank_row_78"])

    def test_normalize_known_summary_rows_excludes_sigma_from_data(self):
        proposal = TableBoundaryProposal(
            proposal_id="proposal_sigma",
            candidate_id="candidate_sigma",
            sheet_name="ban1_%Sig",
            source_range="A2:AB77",
            regions=TableRegions(header_rows=[10, 11], data_rows=[14, 77], footnote_rows=[4, 5]),
        )
        path = Path(__file__).parents[2] / "outputs/ark_smoke/Tabs_%95_first_table.xlsx"
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            normalized, corrections = normalize_known_summary_rows(proposal, workbook["ban1_%Sig"])
        finally:
            workbook.close()
        self.assertEqual(normalized.regions.data_rows, [14])
        self.assertEqual(normalized.regions.footnote_rows, [4, 5, 77])
        self.assertEqual(corrections, ["reclassified_sigma_row_77_as_footnote"])
